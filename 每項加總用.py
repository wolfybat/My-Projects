from tkinter import *
from openpyxl import Workbook
from openpyxl import load_workbook
import tkinter.font as font
from tkinter import filedialog, Text



root = Tk()
root.title('Excel_SUM')
root.geometry("700x700")
ws = None
wb = None
filename = ''

# Font
myFont = font.Font(size=30)
entryFont = font.Font(size=15)


# Button functions
def GetSum():
    targetCol_1 = firstCol.get()
    targetCol_2 = secCol.get()
    saveCol_1 = saveColFirst.get()
    saveCol_2 = saveColSec.get()

    cols = len(ws['1'])
    rows = len(ws[targetCol_1])

    for i in range(1, rows + 1):
        excelArr.append([ws[f"{targetCol_1}{i}"].value, ws[f"{targetCol_2}{i}"].value])
    
    newArr = []
    temp = ''
    sum = 0
    for i in range(rows):
        if temp != excelArr[i][0]:
            temp = excelArr[i][0]
            Sum = excelArr[i][1]
            newArr.append([temp, Sum])
        else:
            temp = excelArr[i][0]
            Sum += excelArr[i][1]
            newArr[-1][1] = Sum

    newRows = len(newArr)
    for i in range(newRows):
        ws[f'{saveCol_1}{i + 1}'] = newArr[i][0]
        ws[f'{saveCol_2}{i + 1}'] = newArr[i][1]

    wb.save(f'{filename}')

def Browse():
    global filename
    filename = filedialog.askopenfilename(initialdir="/", title="選擇檔案", filetypes=(("Excel 檔", "*.xlsx"), ))
    fileName.config(text=filename)

def Enter():
    global ws
    global wb
    
    # Create workbook instance
    wb = Workbook()
    
    # Load existing workbook
    wb = load_workbook(f'{filename}')

    # Create active worksheet
    ws = wb.active



excelArr = []

    
# Create button
DoButton = Button(root, text="2. 輸出各項加總", command=GetSum, height=5, width=20, font=myFont)
browseButton = Button(root, text="瀏覽", command=Browse, font=myFont)
enterButton = Button(root, text="1. 確定", command=Enter, font=myFont)
DoButton.pack(pady=30)

# Create input box
filenameLabel = Label(root, text="檔案名稱 :  ", font=entryFont)
fileName = Label(root, text=" ", font=entryFont)
firstColLabel = Label(root, text="第一欄位置 :  ", font=entryFont)
secColLabel = Label(root, text="第二欄位置 :  ", font=entryFont)
firstCol = Entry(root, font=entryFont)
secCol = Entry(root, font=entryFont)
saveColFirstLabel = Label(root, text="存去的第一欄位置 :  ", font=entryFont)
saveColFirst = Entry(root, font=entryFont)
saveColSecLabel = Label(root, text="存去的第二欄位置 :  ", font=entryFont)
saveColSec = Entry(root, font=entryFont)

# Pack widgets
browseButton.pack()
filenameLabel.pack()
fileName.pack()
firstColLabel.pack()
firstCol.pack()
secColLabel.pack()
secCol.pack()
saveColFirstLabel.pack()
saveColFirst.pack()
saveColSecLabel.pack()
saveColSec.pack()
enterButton.pack()


root.mainloop()
